import os
import csv
import io
import uuid
import json
import re
from datetime import datetime
from typing import Any, Dict, List, Tuple

import psycopg2
from psycopg2.extras import execute_values
from dotenv import load_dotenv

from fastapi import FastAPI, UploadFile, File, Query, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

# pip install openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# -----------------------------
# Config
# -----------------------------
load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL missing. Put it in .env")

# Schema-qualified tables
SCHEMA = "datasets"
DATASETS_TBL = f"{SCHEMA}.datasets"
ROWS_TBL = f"{SCHEMA}.dataset_rows"
STATS_TBL = f"{SCHEMA}.dataset_column_stats"
USAGE_TBL = f"{SCHEMA}.dataset_filter_usage"

app = FastAPI(title="AIAG-07 Backend (Upload CSV + DataTables + Agents + Export)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -----------------------------
# DB
# -----------------------------
def get_conn():
    return psycopg2.connect(DATABASE_URL)

# -----------------------------
# Agents
# -----------------------------
def performance_strategy_agent(
    rows: int,
    columns: int,
    text_columns: int,
    used_fts: bool,
    viewport_width: int
) -> Dict[str, Any]:
    """
    Performance Strategy Decision Agent (PSDA)
    """
    strategy = {
        "processingMode": "server",      # client | server
        "renderMode": "pagination",      # pagination | infinite_scroll
        "pageSize": 25,
        "reason": []
    }

    # Processing mode
    if rows <= 5000 and columns <= 12:
        strategy["processingMode"] = "client"
        strategy["reason"].append("Small dataset")

    # Rendering strategy
    if rows > 100000:
        strategy["renderMode"] = "infinite_scroll"
        strategy["pageSize"] = 50
        strategy["reason"].append("Very large dataset")

    # Search-heavy datasets
    if used_fts:
        strategy["processingMode"] = "server"
        strategy["reason"].append("FTS enabled")

    # Mobile
    if viewport_width and viewport_width < 600:
        strategy["pageSize"] = 25
        strategy["reason"].append("Small viewport")

    # Wide screens
    if viewport_width and viewport_width > 1200 and rows < 100000:
        strategy["pageSize"] = 100
        strategy["reason"].append("Large screen")

    return strategy

# -----------------------------
# CSV/Type helpers
# -----------------------------
def infer_type(values: List[str]) -> str:
    """
    Simple inference: int -> float -> datetime -> text
    """
    cleaned = [v for v in values if v is not None and str(v).strip() != ""]
    if not cleaned:
        return "text"

    try:
        for v in cleaned[:200]:
            int(v)
        return "int"
    except Exception:
        pass

    try:
        for v in cleaned[:200]:
            float(v)
        return "float"
    except Exception:
        pass

    dt_formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ]

    def can_parse(s: str) -> bool:
        s = str(s).strip()
        for fmt in dt_formats:
            try:
                datetime.strptime(s, fmt)
                return True
            except Exception:
                continue
        return False

    sample = [v for v in cleaned[:100] if str(v).strip() != ""]
    if sample and all(can_parse(v) for v in sample):
        return "datetime"

    return "text"

def build_search_text(row: Dict[str, Any], cols: List[str]) -> str:
    return " ".join(str(row.get(c, "") or "") for c in cols)

def safe_colname(col: str, allowed_cols: List[str]) -> str:
    return col if col in allowed_cols else (allowed_cols[0] if allowed_cols else "")

# -----------------------------
# Adaptive Filtering Agent infra
# -----------------------------
def bump_filter_usage(conn, dataset_id: str, cols_used: List[str]):
    if not cols_used:
        return
    cur = conn.cursor()
    for col in cols_used:
        cur.execute(
            f"""
            INSERT INTO {USAGE_TBL}(dataset_id, col_name, used_count, last_used_at)
            VALUES (%s,%s,1,now())
            ON CONFLICT (dataset_id, col_name)
            DO UPDATE SET
              used_count = {USAGE_TBL}.used_count + 1,
              last_used_at = now()
            """,
            (dataset_id, col)
        )

def compute_and_store_column_stats(conn, dataset_id: str, allowed_cols: List[str], col_types: Dict[str, str]):
    """
    Compute lightweight stats:
    - distinct_count, null_count, total_count
    - top_values (for enum-like columns)
    Stored in datasets.dataset_column_stats
    """
    cur = conn.cursor()

    cur.execute(f"SELECT COUNT(*) FROM {ROWS_TBL} WHERE dataset_id=%s", (dataset_id,))
    total_count = cur.fetchone()[0] or 0

    for col in allowed_cols:
        # Null/empty count
        cur.execute(
            f"""
            SELECT COUNT(*)
            FROM {ROWS_TBL}
            WHERE dataset_id=%s AND (row_json->>%s IS NULL OR btrim(row_json->>%s) = '')
            """,
            (dataset_id, col, col)
        )
        null_count = cur.fetchone()[0] or 0

        # Distinct count (non-empty)
        cur.execute(
            f"""
            SELECT COUNT(DISTINCT row_json->>%s)
            FROM {ROWS_TBL}
            WHERE dataset_id=%s AND btrim(COALESCE(row_json->>%s,'')) <> ''
            """,
            (col, dataset_id, col)
        )
        distinct_count = cur.fetchone()[0] or 0

        # Top values if enum-ish
        top_values: List[str] = []
        if distinct_count <= 50 and (col_types.get(col, "text") == "text"):
            cur.execute(
                f"""
                SELECT row_json->>%s AS v, COUNT(*) AS c
                FROM {ROWS_TBL}
                WHERE dataset_id=%s AND btrim(COALESCE(row_json->>%s,'')) <> ''
                GROUP BY v
                ORDER BY c DESC
                LIMIT 15
                """,
                (col, dataset_id, col)
            )
            top_values = [r[0] for r in cur.fetchall() if r[0] is not None]

        cur.execute(
            f"""
            INSERT INTO {STATS_TBL}(dataset_id, col_name, col_type, distinct_count, null_count, total_count, top_values)
            VALUES (%s,%s,%s,%s,%s,%s,%s::jsonb)
            ON CONFLICT (dataset_id, col_name)
            DO UPDATE SET
              col_type=EXCLUDED.col_type,
              distinct_count=EXCLUDED.distinct_count,
              null_count=EXCLUDED.null_count,
              total_count=EXCLUDED.total_count,
              top_values=EXCLUDED.top_values,
              updated_at=now()
            """,
            (dataset_id, col, col_types.get(col, "text"), distinct_count, null_count, total_count, json.dumps(top_values))
        )

def adaptive_filtering_agent(conn, dataset_id: str) -> Dict[str, Any]:
    """
    Suggest best filters using:
    - column stats (distinct_count, top_values, types)
    - usage patterns (used_count)
    """
    cur = conn.cursor()

    cur.execute(
        f"""
        SELECT col_name, col_type, distinct_count, null_count, total_count, top_values
        FROM {STATS_TBL}
        WHERE dataset_id=%s
        """,
        (dataset_id,)
    )
    stats = cur.fetchall()

    cur.execute(
        f"""
        SELECT col_name, used_count
        FROM {USAGE_TBL}
        WHERE dataset_id=%s
        """,
        (dataset_id,)
    )
    usage = {r[0]: r[1] for r in cur.fetchall()}

    suggestions = []

    for col_name, col_type, distinct_count, null_count, total_count, top_values in stats:
        used_count = usage.get(col_name, 0)
        non_null_ratio = 0 if not total_count else (1 - (null_count / total_count))

        score = 0
        score += used_count * 3
        score += 8 if col_type == "datetime" else 0
        score += 6 if col_type in ("int", "float") else 0
        score += 8 if distinct_count is not None and distinct_count <= 20 else 0
        score += 5 if non_null_ratio > 0.7 else 0

        ui_type = "text"
        config: Dict[str, Any] = {}

        if col_type == "datetime":
            ui_type = "date_range"
        elif col_type in ("int", "float"):
            ui_type = "number_range"
        elif distinct_count is not None and distinct_count <= 20 and top_values:
            ui_type = "enum"
            config["options"] = top_values

        suggestions.append({
            "column": col_name,
            "type": ui_type,
            "score": score,
            "config": config,
            "usedCount": used_count,
            "distinctCount": distinct_count
        })

    suggestions.sort(key=lambda x: x["score"], reverse=True)
    top = suggestions[:6]
    combo = [f["column"] for f in top if f["type"] in ("enum", "date_range")][:3]

    return {"recommendedFilters": top, "recommendedCombo": combo}

# -----------------------------
# Export helpers
# -----------------------------
def sanitize_filename(name: str) -> str:
    name = name or "export"
    name = re.sub(r"[^a-zA-Z0-9._-]+", "_", name).strip("_")
    return name[:80] or "export"

def build_where_and_params(
    dataset_id: str,
    allowed_cols: List[str],
    search_value: str,
    status: str,
    category: str,
    fromDate: str,
    toDate: str
) -> Tuple[str, List[Any], bool]:
    where = ["dataset_id = %s"]
    params: List[Any] = [dataset_id]

    if status and "status" in allowed_cols:
        where.append("row_json->>'status' = %s")
        params.append(status)

    if category and "category" in allowed_cols:
        where.append("row_json->>'category' = %s")
        params.append(category)

    if fromDate and "created_at" in allowed_cols:
        where.append("NULLIF(row_json->>'created_at','')::timestamptz >= %s::timestamptz")
        params.append(fromDate)

    if toDate and "created_at" in allowed_cols:
        where.append("NULLIF(row_json->>'created_at','')::timestamptz <= %s::timestamptz")
        params.append(toDate)

    used_fts = False
    if search_value and search_value.strip():
        where.append("search_tsv @@ plainto_tsquery('simple', %s)")
        params.append(search_value.strip())
        used_fts = True

    return " AND ".join(where), params, used_fts

# -----------------------------
# Routes
# -----------------------------
@app.get("/")
def home():
    return {"message": "AIAG-07 backend running. Go to /docs"}

# -----------------------------
# Upload CSV
# -----------------------------
@app.post("/api/datasets/upload")
async def upload_dataset(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode("utf-8", errors="ignore")

    reader = csv.DictReader(io.StringIO(text))
    columns = reader.fieldnames or []
    if not columns:
        raise HTTPException(status_code=400, detail="CSV must have headers in the first row")

    rows: List[Dict[str, Any]] = []
    sample_values: Dict[str, List[str]] = {c: [] for c in columns}

    for i, row in enumerate(reader):
        rows.append(row)
        if i < 300:
            for c in columns:
                sample_values[c].append(row.get(c, ""))

    col_types = {c: infer_type(sample_values[c]) for c in columns}

    dataset_id = uuid.uuid4()
    dataset_name = file.filename or "dataset.csv"

    conn = get_conn()
    try:
        conn.autocommit = False
        cur = conn.cursor()

        cur.execute(
            f"""
            INSERT INTO {DATASETS_TBL}(dataset_id, name, columns_json, column_types)
            VALUES (%s, %s, %s::jsonb, %s::jsonb)
            """,
            (str(dataset_id), dataset_name, json.dumps(columns), json.dumps(col_types)),
        )

        values = []
        for r in rows:
            search_text = build_search_text(r, columns)
            values.append((str(dataset_id), json.dumps(r), search_text, search_text))

        execute_values(
            cur,
            f"""
            INSERT INTO {ROWS_TBL}(dataset_id, row_json, search_text, search_tsv)
            VALUES %s
            """,
            values,
            template="(%s, %s::jsonb, %s, to_tsvector('simple', %s))"
        )

        # compute stats for adaptive filtering
        compute_and_store_column_stats(conn, str(dataset_id), columns, col_types)

        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
    finally:
        conn.close()

    return {
        "datasetId": str(dataset_id),
        "name": dataset_name,
        "rows": len(rows),
        "columns": columns,
        "columnTypes": col_types
    }

# -----------------------------
# DataTables endpoint
# -----------------------------
@app.get("/api/table")
def datatable_endpoint(
    datasetId: str = Query(...),
    draw: int = Query(1),
    start: int = Query(0),
    length: int = Query(25),
    search_value: str = Query("", alias="search[value]"),
    order_col_index: int = Query(0, alias="order[0][column]"),
    order_dir: str = Query("asc", alias="order[0][dir]"),
    status: str = Query(""),
    category: str = Query(""),
    fromDate: str = Query(""),
    toDate: str = Query(""),
    viewportWidth: int = Query(0),
):
    if length < 1:
        length = 25
    if length > 500:
        length = 500

    conn = get_conn()
    try:
        cur = conn.cursor()

        cur.execute(f"SELECT columns_json, column_types FROM {DATASETS_TBL} WHERE dataset_id=%s", (datasetId,))
        meta = cur.fetchone()
        if not meta:
            raise HTTPException(status_code=404, detail="Invalid datasetId")

        allowed_cols: List[str] = meta[0] or []
        col_types: Dict[str, str] = meta[1] or {}

        sort_col = allowed_cols[order_col_index] if (0 <= order_col_index < len(allowed_cols)) else (allowed_cols[0] if allowed_cols else "")
        sort_col = safe_colname(sort_col, allowed_cols)
        sort_dir = "ASC" if str(order_dir).lower() == "asc" else "DESC"

        t = (col_types.get(sort_col) or "text").lower()
        if t in ("int", "float"):
            sort_expr = f"NULLIF((row_json->>'{sort_col}'),'')::numeric"
        elif t == "datetime":
            sort_expr = f"NULLIF((row_json->>'{sort_col}'),'')::timestamptz"
        else:
            sort_expr = f"(row_json->>'{sort_col}')"

        where = ["dataset_id = %s"]
        params: List[Any] = [datasetId]

        if status and "status" in allowed_cols:
            where.append("row_json->>'status' = %s")
            params.append(status)

        if category and "category" in allowed_cols:
            where.append("row_json->>'category' = %s")
            params.append(category)

        if fromDate and "created_at" in allowed_cols:
            where.append("NULLIF(row_json->>'created_at','')::timestamptz >= %s::timestamptz")
            params.append(fromDate)

        if toDate and "created_at" in allowed_cols:
            where.append("NULLIF(row_json->>'created_at','')::timestamptz <= %s::timestamptz")
            params.append(toDate)

        used_fts = False
        if search_value and search_value.strip():
            where.append("search_tsv @@ plainto_tsquery('simple', %s)")
            params.append(search_value.strip())
            used_fts = True

        where_sql = " AND ".join(where)

        # counts
        cur.execute(f"SELECT COUNT(*) FROM {ROWS_TBL} WHERE dataset_id=%s", (datasetId,))
        records_total = cur.fetchone()[0] or 0

        cur.execute(f"SELECT COUNT(*) FROM {ROWS_TBL} WHERE {where_sql}", tuple(params))
        records_filtered = cur.fetchone()[0] or 0

        # page query
        query = f"""
            SELECT row_json
            FROM {ROWS_TBL}
            WHERE {where_sql}
            ORDER BY {sort_expr} {sort_dir} NULLS LAST
            LIMIT %s OFFSET %s
        """
        params_page = params + [int(length), int(start)]
        cur.execute(query, tuple(params_page))
        rows = cur.fetchall()

        data = [r[0] for r in rows]
        text_cols = sum(1 for _t in col_types.values() if str(_t).lower() == "text")

        agent_strategy = performance_strategy_agent(
            rows=records_total,
            columns=len(allowed_cols),
            text_columns=text_cols,
            used_fts=used_fts,
            viewport_width=viewportWidth
        )

        # usage tracking
        cols_used = []
        if status and "status" in allowed_cols:
            cols_used.append("status")
        if category and "category" in allowed_cols:
            cols_used.append("category")
        if (fromDate or toDate) and "created_at" in allowed_cols:
            cols_used.append("created_at")
        if search_value and search_value.strip():
            cols_used.append("__global_search__")
        bump_filter_usage(conn, datasetId, cols_used)

        # adaptive filters
        adaptive_filters = adaptive_filtering_agent(conn, datasetId)

        agent_insights = {
            "mode": "server",
            "usedFTS": used_fts,
            "datasetRows": records_total,
            "filteredRows": records_filtered,
            "sortColumn": sort_col,
            "sortDir": sort_dir,
            "viewportWidth": viewportWidth,
            "agentStrategy": agent_strategy,
            "adaptiveFilters": adaptive_filters
        }

        return {
            "draw": draw,
            "recordsTotal": records_total,
            "recordsFiltered": records_filtered,
            "data": data,
            "agentInsights": agent_insights
        }

    finally:
        conn.close()

# -----------------------------
# Export CSV
# -----------------------------
@app.get("/api/export/csv")
def export_csv(
    datasetId: str = Query(...),
    status: str = Query(""),
    category: str = Query(""),
    fromDate: str = Query(""),
    toDate: str = Query(""),
    search_value: str = Query("", alias="search[value]"),
):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(f"SELECT name, columns_json FROM {DATASETS_TBL} WHERE dataset_id=%s", (datasetId,))
    meta = cur.fetchone()
    if not meta:
        conn.close()
        raise HTTPException(status_code=404, detail="Invalid datasetId")

    dataset_name, allowed_cols = meta[0], (meta[1] or [])

    where_sql, params, used_fts = build_where_and_params(
        dataset_id=datasetId,
        allowed_cols=allowed_cols,
        search_value=search_value,
        status=status,
        category=category,
        fromDate=fromDate,
        toDate=toDate
    )

    filename = sanitize_filename(dataset_name.replace(".csv", "")) + "_export.csv"

    def row_gen():
        # header
        yield ",".join([f'"{c.replace("\"","\"\"")}"' for c in allowed_cols]) + "\n"

        export_cur = conn.cursor(name="export_csv_cursor")
        export_cur.itersize = 2000

        export_cur.execute(
            f"SELECT row_json FROM {ROWS_TBL} WHERE {where_sql} ORDER BY id ASC",
            tuple(params),
        )

        for (row_json,) in export_cur:
            vals = []
            for c in allowed_cols:
                v = row_json.get(c, "")
                if v is None:
                    v = ""
                v = str(v).replace('"', '""')
                vals.append(f'"{v}"')
            yield ",".join(vals) + "\n"

        export_cur.close()
        conn.close()

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Used-FTS": "true" if used_fts else "false"
    }

    return StreamingResponse(row_gen(), media_type="text/csv", headers=headers)

# -----------------------------
# Export Excel
# -----------------------------
@app.get("/api/export/excel")
def export_excel(
    datasetId: str = Query(...),
    status: str = Query(""),
    category: str = Query(""),
    fromDate: str = Query(""),
    toDate: str = Query(""),
    search_value: str = Query("", alias="search[value]"),
):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(f"SELECT name, columns_json FROM {DATASETS_TBL} WHERE dataset_id=%s", (datasetId,))
    meta = cur.fetchone()
    if not meta:
        conn.close()
        raise HTTPException(status_code=404, detail="Invalid datasetId")

    dataset_name, allowed_cols = meta[0], (meta[1] or [])

    where_sql, params, used_fts = build_where_and_params(
        dataset_id=datasetId,
        allowed_cols=allowed_cols,
        search_value=search_value,
        status=status,
        category=category,
        fromDate=fromDate,
        toDate=toDate
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(allowed_cols)

    export_cur = conn.cursor(name="export_excel_cursor")
    export_cur.itersize = 2000

    export_cur.execute(
        f"SELECT row_json FROM {ROWS_TBL} WHERE {where_sql} ORDER BY id ASC",
        tuple(params),
    )

    for (row_json,) in export_cur:
        ws.append([row_json.get(c, "") for c in allowed_cols])

    export_cur.close()
    conn.close()

    for i, col in enumerate(allowed_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(col) + 2, 12), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = sanitize_filename(dataset_name.replace(".csv", "")) + "_export.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Used-FTS": "true" if used_fts else "false"
    }

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
